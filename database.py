"""
Gerenciamento de dados via arquivo JSON
"""

import json
import os
import csv
import getpass
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from models import Aula, Laboratorio
from config import (
    DATABASE_PATH
)
from utils import formatar_cor_hex

# Ordem dos dias para ordenação
_ORDEM_DIAS = {
    "SEGUNDA-FEIRA": 1,
    "TERÇA-FEIRA": 2,
    "QUARTA-FEIRA": 3,
    "QUINTA-FEIRA": 4,
    "SEXTA-FEIRA": 5,
    "SÁBADO": 6,
}

# Estrutura padrão do JSON
_ESTRUTURA_VAZIA = {
    "proximo_id": 1,
    "aulas": [],
    "faculdades": [],
    "cursos_por_faculdade": {},
    "cores_cursos": {},
    "disciplinas": [], # lista global (cache/retro compatibilidade)
    "turmas": [],      # lista global (cache/retro compatibilidade)
    "disciplinas_por_curso": {},
    "turmas_por_disciplina": {},
    "alunos_por_turma": {}, # { "DISCIPLINA|TURMA": qtde }
    "laboratorios": [],
    "status_laboratorios": {}, # { "Lab 01": "Ligado", ... }
    "status_acoes_labs": {},   # { "Lab 01": "Organizar", ... }
    "status_horario": "8h",
    "historico": [], # [{ "timestamp": "...", "acao": "...", "detalhes": "..." }, ...]
}


class Database:
    def __init__(self):
        self._path = DATABASE_PATH
        self._dados = self._carregar()
        self._semear_laboratorios()

    def _semear_laboratorios(self):
        """Cria laboratórios de 1 a 14 se não existirem."""
        existentes = {l.nome for l in self.listar_laboratorios()}
        alterado = False
        for i in range(1, 15):
            nome_lab = f"Laboratório {i}"
            if nome_lab not in existentes:
                lab = Laboratorio(
                    id=None,
                    nome=nome_lab,
                    descricao=f"Descrição padrão para o {nome_lab}",
                    observacao="",
                    qtd_micros=0,
                    planta_path=""
                )
                self.adicionar_laboratorio(lab)
                alterado = True
        
        if alterado:
            self._salvar()

    # ------------------------------------------------------------------
    # Persistência
    # ------------------------------------------------------------------

    def _carregar(self) -> dict:
        """Lê o arquivo JSON ou retorna estrutura vazia."""
        dados = self._carregar_arquivo()
        return dados if dados is not None else dict(_ESTRUTURA_VAZIA)

    def _carregar_arquivo(self) -> Optional[dict]:
        """Lê o arquivo JSON do disco. Retorna None se houver erro (exceto se o arquivo não existir)."""
        if not os.path.exists(self._path):
            return dict(_ESTRUTURA_VAZIA)
            
        try:
            with open(self._path, "r", encoding="utf-8") as f:
                carregado = json.load(f)
            
            dados = dict(_ESTRUTURA_VAZIA)
            dados.update(carregado)
            # Garante que chaves novas existam em arquivos antigos
            for chave, valor in _ESTRUTURA_VAZIA.items():
                if chave not in dados:
                    dados[chave] = valor
            return dados
        except (json.JSONDecodeError, OSError):
            return None

    # Método interno para limpar qualquer lixo ou espaços de chaves no dicionário
    def _normalizar_base(self):
        """Standardiza todas as chaves de cursos e faculdades no JSON."""
        # 1. Cores
        if "cores_cursos" in self._dados:
            novas_cores = {}
            for k, v in self._dados["cores_cursos"].items():
                if k: novas_cores[k.strip().upper()] = v
            self._dados["cores_cursos"] = novas_cores
            
        # 2. Cursos por Faculdade
        if "cursos_por_faculdade" in self._dados:
            nova_cpf = {}
            for fac, cursos in self._dados["cursos_por_faculdade"].items():
                fac_n = fac.strip().upper()
                nova_cpf[fac_n] = [c.strip().upper() for c in cursos if c]
            self._dados["cursos_por_faculdade"] = nova_cpf

        # 3. Disciplinas por Curso
        if "disciplinas_por_curso" in self._dados:
            nova_dpc = {}
            for curso, discs in self._dados["disciplinas_por_curso"].items():
                curso_n = curso.strip().upper()
                nova_dpc[curso_n] = discs
            self._dados["disciplinas_por_curso"] = nova_dpc

    def _salvar(self):
        """Salva os dados no arquivo JSON com normalização prévia e gravação atômica."""
        self._normalizar_base()
        
        # Gera o JSON em memória primeiro
        conteudo = json.dumps(self._dados, ensure_ascii=False, indent=None) # indent=None para performance e menor tamanho
        
        # Gravação Atômica: Salva em um arquivo temporário e depois renomeia.
        # Isso evita que o arquivo original seja corrompido em caso de falha de rede ou queda de energia.
        temp_path = self._path + ".tmp"
        try:
            with open(temp_path, "w", encoding="utf-8") as f:
                f.write(conteudo)
            
            # No Windows, os.replace substitui o arquivo de forma atômica se estiver no mesmo volume
            if os.path.exists(self._path):
                os.replace(temp_path, self._path)
            else:
                os.rename(temp_path, self._path)
        except Exception as e:
            if os.path.exists(temp_path):
                try: os.remove(temp_path)
                except: pass
            raise e

    def _dict_para_aula(self, d: dict) -> Aula:
        curso = d.get("curso", "")
        # Fallback: se a aula não tem cor ou é branca, tenta pegar a cor do curso
        cor = d.get("cor_fundo") or "#ffffff"
        if cor == "#ffffff" and curso:
            cor = self.obter_cor_curso(curso)
        
        cor = formatar_cor_hex(cor)
            
        return Aula(
            id=d["id"],
            laboratorio=d["laboratorio"],
            dia_semana=d["dia_semana"],
            hora_inicio=d["hora_inicio"],
            hora_fim=d["hora_fim"],
            disciplina=d["disciplina"],
            turma=d["turma"],
            professor=d.get("professor"),
            qtde_alunos=d.get("qtde_alunos"),
            faculdade=d["faculdade"],
            curso=curso,
            cor_fundo=cor,
            observacoes=d.get("observacoes", ""),
            peso_observacao=d.get("peso_observacao", "Baixo"),
            is_eventual=d.get("is_eventual", False),
            data_eventual=d.get("data_eventual")
        )

    @staticmethod
    def _aula_para_dict(aula: Aula) -> dict:
        return {
            "id": aula.id,
            "laboratorio": aula.laboratorio,
            "dia_semana": aula.dia_semana,
            "hora_inicio": aula.hora_inicio,
            "hora_fim": aula.hora_fim,
            "disciplina": aula.disciplina,
            "turma": aula.turma,
            "professor": aula.professor,
            "qtde_alunos": aula.qtde_alunos,
            "faculdade": aula.faculdade,
            "curso": aula.curso,
            "cor_fundo": aula.cor_fundo,
            "observacoes": aula.observacoes,
            "peso_observacao": aula.peso_observacao,
            "is_eventual": aula.is_eventual,
            "data_eventual": aula.data_eventual
        }

    def _dict_para_laboratorio(self, d: dict) -> Laboratorio:
        return Laboratorio(
            id=d["id"],
            nome=d["nome"],
            descricao=d["descricao"],
            observacao=d["observacao"],
            qtd_micros=d.get("qtd_micros", 0),
            planta_path=d.get("planta_path"),
            inventario=d.get("inventario", {})
        )

    @staticmethod
    def _laboratorio_para_dict(lab: Laboratorio) -> dict:
        return {
            "id": lab.id,
            "nome": lab.nome,
            "descricao": lab.descricao,
            "observacao": lab.observacao,
            "qtd_micros": lab.qtd_micros,
            "planta_path": lab.planta_path,
            "inventario": lab.inventario if lab.inventario is not None else {}
        }

    # ------------------------------------------------------------------
    # CRUD — Aulas
    # ------------------------------------------------------------------

    def adicionar_aula(self, aula: Aula) -> int:
        """Adiciona uma nova aula e retorna o ID gerado."""
        novo_id = self._dados["proximo_id"]
        self._dados["proximo_id"] += 1
        aula.id = novo_id
        self._dados["aulas"].append(self._aula_para_dict(aula))
        
        self.registrar_log("INSERÇÃO", f"Aula ID {aula.id}: {aula.disciplina} ({aula.laboratorio} - {aula.dia_semana} {aula.hora_inicio}-{aula.hora_fim})")
        
        self._salvar()
        return novo_id

    def alterar_aula(self, aula: Aula):
        """Altera uma aula existente pelo ID, registrando detalhadamente as mudanças."""
        for i, d in enumerate(self._dados["aulas"]):
            if d["id"] == aula.id:
                # Obter dados antigos para comparação
                antiga = self._dict_para_aula(d)
                mudancas = []
                
                # Mapeamento de atributos para labels amigáveis no log
                campos = [
                    ("laboratorio", "Lab"),
                    ("dia_semana", "Dia"),
                    ("hora_inicio", "Ini"),
                    ("hora_fim", "Fim"),
                    ("disciplina", "Disc"),
                    ("turma", "Turma"),
                    ("professor", "Prof"),
                    ("faculdade", "Fac"),
                    ("curso", "Curso"),
                    ("qtde_alunos", "Qtd"),
                    ("cor_fundo", "Cor"),
                    ("observacoes", "Obs"),
                    ("peso_observacao", "Peso_Obs"),
                    ("is_eventual", "Eventual"),
                    ("data_eventual", "Data_Evt")
                ]
                
                for attr, label in campos:
                    v_old = getattr(antiga, attr)
                    v_new = getattr(aula, attr)
                    # Compara strings ou valores básicos
                    if str(v_old).strip().upper() != str(v_new).strip().upper():
                        mudancas.append(f"{label}: {v_old} → {v_new}")
                
                if mudancas:
                    detalhes = f"ID {aula.id} ({aula.disciplina}): " + " | ".join(mudancas)
                else:
                    detalhes = f"ID {aula.id} ({aula.disciplina}): Atualizado (sem mudanças visíveis nos campos principais)"

                self._dados["aulas"][i] = self._aula_para_dict(aula)
                self.registrar_log("ALTERAÇÃO", detalhes)
                self._salvar()
                return

    def apagar_aula(self, aula_id: int):
        """Remove uma aula pelo ID."""
        for d in self._dados["aulas"]:
            if d["id"] == aula_id:
                aula = self._dict_para_aula(d)
                self.registrar_log("EXCLUSÃO", f"Aula ID {aula.id}: {aula.disciplina} ({aula.laboratorio} - {aula.dia_semana} {aula.hora_inicio}-{aula.hora_fim})")
                break

        self._dados["aulas"] = [
            d for d in self._dados["aulas"] if d["id"] != aula_id
        ]
        self._salvar()

    def apagar_aulas_lote(self, lista_ids: List[int]):
        """Remove múltiplas aulas de uma vez."""
        ids_set = set(lista_ids)
        
        # Registrar logs antes de apagar para ter os detalhes
        for d in self._dados["aulas"]:
            if d["id"] in ids_set:
                aula = self._dict_para_aula(d)
                self.registrar_log("EXCLUSÃO", f"Aula ID {aula.id}: {aula.disciplina} ({aula.laboratorio} - {aula.dia_semana} {aula.hora_inicio}-{aula.hora_fim})")

        self._dados["aulas"] = [
            d for d in self._dados["aulas"] if d["id"] not in ids_set
        ]
        self._salvar()

    # ------------------------------------------------------------------
    # Histórico / Logs
    # ------------------------------------------------------------------

    def registrar_log(self, acao: str, detalhes: str):
        """Adiciona uma entrada ao histórico de modificações incluindo o usuário do sistema."""
        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        try:
            # Captura o usuário logado no Windows/Sistema
            usuario = getpass.getuser().upper()
        except:
            usuario = "DESCONHECIDO"

        log = {
            "timestamp": agora,
            "usuario": usuario,
            "acao": acao,
            "detalhes": detalhes
        }
        self._dados.setdefault("historico", []).insert(0, log) # Mais recentes primeiro
        # Limitar histórico a 1000 entradas para não sobrecarregar o JSON
        if len(self._dados["historico"]) > 1000:
            self._dados["historico"] = self._dados["historico"][:1000]
        self._salvar()

    def listar_historico(self) -> List[dict]:
        """Retorna a lista completa de logs do histórico."""
        return self._dados.get("historico", [])

    def limpar_historico(self):
        """Remove todos os registros do histórico."""
        self._dados["historico"] = []
        self._salvar()

    # ------------------------------------------------------------------
    # CRUD — Laboratórios
    # ------------------------------------------------------------------

    def adicionar_laboratorio(self, lab: Laboratorio) -> int:
        """Adiciona um novo laboratório."""
        # Se 'laboratorios' ainda não existir, inicializa (garantido pelo _carregar/vazia mas por segurança)
        if "laboratorios" not in self._dados:
            self._dados["laboratorios"] = []
            
        # Reutilizar proximo_id ou criar um novo contador? 
        # O proximo_id global parece ser usado apenas para aulas.
        # Vamos usar um contador separado para labs no JSON se preferir, ou apenas o global.
        # Para evitar conflitos, vou usar o global mas incrementar.
        novo_id = self._dados["proximo_id"]
        self._dados["proximo_id"] += 1
        lab.id = novo_id
        self._dados["laboratorios"].append(self._laboratorio_para_dict(lab))
        self._salvar()
        return novo_id

    def alterar_laboratorio(self, lab: Laboratorio):
        """Altera um laboratório existente."""
        for i, d in enumerate(self._dados.get("laboratorios", [])):
            if d["id"] == lab.id:
                self._dados["laboratorios"][i] = self._laboratorio_para_dict(lab)
                self._salvar()
                return

    def apagar_laboratorio(self, lab_id: int):
        """Remove um laboratório pelo ID."""
        self._dados["laboratorios"] = [
            d for d in self._dados.get("laboratorios", []) if d["id"] != lab_id
        ]
        self._salvar()

    def listar_laboratorios(self) -> List[Laboratorio]:
        """Lista todos os laboratórios cadastrados."""
        return [
            self._dict_para_laboratorio(d) 
            for d in self._dados.get("laboratorios", [])
        ]

    # ------------------------------------------------------------------
    # Consultas — Aulas
    # ------------------------------------------------------------------

    def listar_todas_aulas(self) -> List[Aula]:
        """Lista todas as aulas ordenadas por laboratório, dia e hora."""
        ordenadas = sorted(
            self._dados["aulas"],
            key=lambda d: (
                str(d["laboratorio"]).zfill(5) if str(d["laboratorio"]).isdigit() else str(d["laboratorio"]),
                _ORDEM_DIAS.get(str(d["dia_semana"]).upper(), 99),
                str(d["hora_inicio"]),
            ),
        )
        res = []
        for d in ordenadas:
            aula = self._dict_para_aula(d)
            # Fallback: Se não tem cor ou é branca, tenta pegar a cor atual do curso
            cor_atual = d.get("cor_fundo", "#ffffff")
            if cor_atual == "#ffffff" or not cor_atual:
                aula.cor_fundo = self.obter_cor_curso(aula.curso)
            res.append(aula)
        return res

    def obter_estatisticas_aulas(self) -> dict:
        """Retorna contagens de diferentes tipos de aulas para relatórios."""
        aulas = self._dados.get("aulas", [])
        
        semestral = 0
        eventual = 0
        
        for a in aulas:
            if a.get("is_eventual", False):
                eventual += 1
            else:
                semestral += 1
                
        return {
            "semestral": semestral,
            "eventual": eventual,
            "reposicao": None, # Ainda não implementado
            "cancelada": None,  # Ainda não implementado
            "indeferida": None  # Ainda não implementado
        }

    def verificar_conflito(
        self, aula: Aula, ignorar_id: Optional[int] = None
    ) -> Optional[Aula]:
        """Verifica conflito de horário. Retorna a aula conflitante ou None.

        Regras:
        - Aula eventual sendo cadastrada: nunca conflita (pode sobrepor fixas/eventuais).
        - Aula fixa sendo cadastrada: conflita apenas com outras aulas FIXAS no mesmo
          laboratório, dia e horário. Aulas eventuais existentes não bloqueiam o cadastro
          de fixas, pois são exceções pontuais que não ocupam o slot permanentemente.
        """
        # Aulas eventuais podem ser cadastradas em qualquer horário
        if aula.is_eventual:
            return None

        for d in self._dados["aulas"]:
            if ignorar_id is not None and d["id"] == ignorar_id:
                continue
            if d["laboratorio"] != aula.laboratorio:
                continue
            if d["dia_semana"] != aula.dia_semana:
                continue

            # Aulas eventuais existentes não bloqueiam o cadastro de aulas fixas
            if d.get("is_eventual", False):
                continue

            if d["hora_inicio"] < aula.hora_fim and d["hora_fim"] > aula.hora_inicio:
                return self._dict_para_aula(d)
        return None

    def listar_aulas_por_dia_lab(
        self, dia_semana: str, laboratorio: str
    ) -> List[Aula]:
        """Lista aulas de um dia e laboratório específicos."""
        filtradas = [
            d
            for d in self._dados["aulas"]
            if d["dia_semana"] == dia_semana and d["laboratorio"] == laboratorio
        ]
        filtradas.sort(key=lambda d: d["hora_inicio"])
        return [self._dict_para_aula(d) for d in filtradas]

    def listar_aulas_por_lab(self, laboratorio: str) -> List[Aula]:
        """Lista aulas de um laboratório específico (todos os dias)."""
        filtradas = [
            d for d in self._dados["aulas"] if d["laboratorio"] == laboratorio
        ]
        filtradas.sort(
            key=lambda d: (
                _ORDEM_DIAS.get(d["dia_semana"], 99),
                d["hora_inicio"],
            )
        )
        return [self._dict_para_aula(d) for d in filtradas]

    # ------------------------------------------------------------------
    # CRUD — Faculdades
    # ------------------------------------------------------------------

    def listar_faculdades(self) -> List[str]:
        """Retorna faculdades cadastradas no JSON, ordenadas."""
        return sorted(set(self._dados.get("faculdades", [])))

    def adicionar_faculdade(self, nome: str, auto_save: bool = True):
        """Adiciona uma faculdade à lista personalizada."""
        nome = nome.strip().upper()
        if not nome:
            return
        existentes = set(self.listar_faculdades())
        if nome not in existentes:
            self._dados.setdefault("faculdades", []).append(nome)
            if auto_save: self._salvar()

    def editar_faculdade(self, nome_antigo: str, nome_novo: str):
        """Renomeia uma faculdade e atualiza todas as referências."""
        nome_antigo = nome_antigo.strip().upper()
        nome_novo = nome_novo.strip().upper()
        if not nome_novo or nome_antigo == nome_novo:
            return

        # 1. Lista de faculdades
        facs = self._dados.get("faculdades", [])
        for i, f in enumerate(facs):
            if f.strip().upper() == nome_antigo:
                facs[i] = nome_novo
                break

        # 2. Cursos por faculdade
        cpf = self._dados.get("cursos_por_faculdade", {})
        if nome_antigo in cpf:
            cpf[nome_novo] = cpf.pop(nome_antigo)

        # 3. Aulas
        count: int = 0
        for aula in self._dados.get("aulas", []):
            if str(aula.get("faculdade", "")).strip().upper() == nome_antigo:
                aula["faculdade"] = nome_novo
                count += 1
        
        self.registrar_log("ALTERAÇÃO GLOBAL", f"Faculdade renomeada: {nome_antigo} → {nome_novo} (afetou {count} aulas)")
        self._salvar()

    def excluir_faculdade(self, nome: str):
        """Remove uma faculdade e seus cursos vinculados (não remove as aulas)."""
        nome = nome.strip().upper()

        # 1. Lista de faculdades
        self._dados["faculdades"] = [
            f for f in self._dados.get("faculdades", [])
            if f.strip().upper() != nome
        ]

        # 2. Cursos por faculdade
        cpf = self._dados.get("cursos_por_faculdade", {})
        cursos_removidos = cpf.pop(nome, [])

        # 3. Cores e disciplinas dos cursos removidos
        cores = self._dados.get("cores_cursos", {})
        dpc = self._dados.get("disciplinas_por_curso", {})
        tpd = self._dados.get("turmas_por_disciplina", {})
        for curso in cursos_removidos:
            c_upper = curso.strip().upper()
            cores.pop(c_upper, None)
            discs = dpc.pop(c_upper, [])
            for d in discs:
                tpd.pop(d.strip().upper(), None)

        self._salvar()

    # ------------------------------------------------------------------
    # CRUD — Cursos
    # ------------------------------------------------------------------

    def listar_cursos(self, faculdade: str) -> List[str]:
        """Retorna cursos de uma faculdade salvos no JSON."""
        faculdade = faculdade.strip().upper()
        personalizados = self._dados.get("cursos_por_faculdade", {}).get(faculdade, [])
        # Inclui cursos já usados em aulas dessa faculdade (suporte a legados)
        das_aulas = [
            d.get("curso") for d in self._dados.get("aulas", [])
            if str(d.get("faculdade", "")).strip().upper() == faculdade and d.get("curso")
        ]
        return sorted(set(personalizados) | set(das_aulas))

    def adicionar_curso(self, faculdade: str, nome: str, cor: str = "#ffffff", auto_save: bool = True):
        """Adiciona um curso a uma faculdade com sua cor."""
        nome = nome.strip().upper()
        faculdade = faculdade.strip().upper()
        if not nome or not faculdade:
            return
        cpf = self._dados.setdefault("cursos_por_faculdade", {})
        cpf.setdefault(faculdade, [])
        if nome not in cpf[faculdade]:
            cpf[faculdade].append(nome)
        
        # Preserva cor existente se a nova for o padrão #ffffff
        cores = self._dados.setdefault("cores_cursos", {})
        if nome not in cores or cor != "#ffffff":
            cores[nome] = formatar_cor_hex(cor)
            
        if auto_save: self._salvar()

    def obter_cor_curso(self, curso: str) -> str:
        """Retorna a cor do curso cadastrada no JSON. Padrão: #FFFFFF."""
        if not curso: return "#ffffff"
        curso = curso.strip().upper()
        return self._dados.get("cores_cursos", {}).get(curso, "#ffffff")

    def obter_todos_cursos(self) -> List[dict]:
        """Retorna lista de todos os cursos: [{'nome': '...', 'faculdade': '...', 'cor': '...'}]"""
        resultado = []
        cpf = self._dados.get("cursos_por_faculdade", {})
        cores = self._dados.get("cores_cursos", {})
        
        # Cria um mapa auxiliar de cores em uppercase para busca robusta
        mapa_cores = {k.strip().upper(): v for k, v in cores.items()}
        
        for fac, cursos in cpf.items():
            for c in cursos:
                nome_upper = c.strip().upper()
                resultado.append({
                    "nome": c,
                    "faculdade": fac,
                    "cor": mapa_cores.get(nome_upper, "#ffffff")
                })
        return sorted(resultado, key=lambda x: (x["faculdade"], x["nome"]))

    def editar_curso(self, nome_antigo: str, nome_novo: str, faculdade: str, nova_cor: str):
        """Atualiza nome e cor de um curso em toda a base."""
        nome_antigo = nome_antigo.strip().upper()
        nome_novo = nome_novo.strip().upper()
        faculdade = faculdade.strip().upper()
        
        if not nome_novo: return
        
        # 1. Atualiza no mapa hierárquico
        cpf = self._dados.get("cursos_por_faculdade", {})
        # Tenta encontrar a faculdade mesmo com casing diferente nas chaves (caso raro no JSON manual)
        target_fac_key = None
        for k in cpf.keys():
            if k.strip().upper() == faculdade:
                target_fac_key = k
                break
        
        if target_fac_key:
            lista_cursos = cpf[target_fac_key]
            for i, c in enumerate(lista_cursos):
                if c.strip().upper() == nome_antigo:
                    lista_cursos[i] = nome_novo
                    break
            
        # 2. Atualiza a cor (sempre atualiza pela chave do nome novo)
        cores = self._dados.setdefault("cores_cursos", {})
        
        # Limpeza robusta: remove qualquer variação do nome antigo no dicionário de cores
        chaves_para_remover = [k for k in cores.keys() if k.strip().upper() == nome_antigo]
        for k in chaves_para_remover:
            del cores[k]
            
        cores[nome_novo] = formatar_cor_hex(nova_cor)
        
        # 3. Atualiza referências em Disciplinas (se houver)
        dpc = self._dados.get("disciplinas_por_curso", {})
        chaves_dpc = [k for k in dpc.keys() if k.strip().upper() == nome_antigo]
        for k in chaves_dpc:
            if nome_novo not in dpc:
                dpc[nome_novo] = dpc.pop(k)
            else:
                dpc[nome_novo].extend(dpc.pop(k))
                dpc[nome_novo] = list(set(dpc[nome_novo])) # uniq
            
        # 4. Atualiza referências em Aulas
        count: int = 0
        for aula in self._dados.get("aulas", []):
            curso_aula = str(aula.get("curso", "")).strip().upper()
            fac_aula = str(aula.get("faculdade", "")).strip().upper()
            
            match_fac = (not faculdade or fac_aula == faculdade)
            match_curso = (curso_aula == nome_antigo)
            
            if match_fac and match_curso:
                aula["curso"] = nome_novo
                aula["cor_fundo"] = formatar_cor_hex(nova_cor)
                count += 1
                
        self.registrar_log("ALTERAÇÃO GLOBAL", f"Curso modificado: {nome_antigo} em {faculdade} → {nome_novo} (afetou {count} aulas)")
        self._salvar()

    def excluir_curso(self, nome: str, faculdade: str):
        """Remove um curso e suas referências (disciplinas, turmas)."""
        nome = nome.upper()
        faculdade = faculdade.upper()
        
        # 1. Remove do mapa hierárquico
        cpf = self._dados.get("cursos_por_faculdade", {})
        if faculdade in cpf and nome in cpf[faculdade]:
            cpf[faculdade].remove(nome)
            
        # 2. Remove a cor
        cores = self._dados.get("cores_cursos", {})
        if nome in cores:
            del cores[nome]
            
        # 3. Remove disciplinas e turmas vinculadas (opcional, mas recomendado para limpeza)
        dpc = self._dados.get("disciplinas_por_curso", {})
        if nome in dpc:
            disciplinas = dpc.pop(nome)
            tpd = self._dados.get("turmas_por_disciplina", {})
            for d in disciplinas:
                if d in tpd:
                    del tpd[d]
                    
        # 4. Nota: Aulas vinculadas a este curso NÃO são removidas automaticamente 
        # para evitar perda de dados acidental, mas ficarão sem curso no UI.
        
        self._salvar()

    # ------------------------------------------------------------------
    # CRUD — Disciplinas
    # ------------------------------------------------------------------

    def listar_todas_disciplinas(self) -> List[str]:
        """Retorna todas as disciplinas (usadas em qualquer curso)."""
        todas = set(self._dados.get("disciplinas", []))
        for curso_dict in self._dados.get("disciplinas_por_curso", {}).values():
            todas.update(curso_dict)
        return sorted(todas)

    def listar_disciplinas(self, curso: str) -> List[str]:
        """Retorna disciplinas relacionadas a um curso."""
        return sorted(self._dados.get("disciplinas_por_curso", {}).get(curso, []))

    def adicionar_disciplina(self, curso: str, nome: str, auto_save: bool = True):
        """Adiciona uma disciplina relacionada a um curso."""
        nome = nome.strip().upper()
        if not nome or not curso:
            return
        
        # Adiciona no mapa hierárquico
        dpc = self._dados.setdefault("disciplinas_por_curso", {})
        dpc.setdefault(curso, [])
        if nome not in dpc[curso]:
            dpc[curso].append(nome)
        
        # Mantém na lista global
        lista = self._dados.setdefault("disciplinas", [])
        if nome not in lista:
            lista.append(nome)
            
        if auto_save: self._salvar()

    # ------------------------------------------------------------------
    # CRUD — Turmas
    # ------------------------------------------------------------------

    def listar_todas_turmas(self) -> List[str]:
        """Retorna todas as turmas (usadas em qualquer disciplina)."""
        todas = set(self._dados.get("turmas", []))
        for disc_dict in self._dados.get("turmas_por_disciplina", {}).values():
            todas.update(disc_dict)
        return sorted(todas)

    def listar_turmas(self, disciplina: str) -> List[str]:
        """Retorna turmas relacionadas a uma disciplina."""
        return sorted(self._dados.get("turmas_por_disciplina", {}).get(disciplina, []))

    def adicionar_turma(self, disciplina: str, nome: str, qtde_alunos: int = 0, auto_save: bool = True):
        """Adiciona uma turma relacionada a uma disciplina."""
        nome = nome.strip().upper()
        disciplina = disciplina.strip().upper()
        if not nome or not disciplina:
            return
            
        # Adiciona no mapa hierárquico
        tpd = self._dados.setdefault("turmas_por_disciplina", {})
        tpd.setdefault(disciplina, [])
        if nome not in tpd[disciplina]:
            tpd[disciplina].append(nome)
            
        # Salva a quantidade de alunos
        apt = self._dados.setdefault("alunos_por_turma", {})
        chave = f"{disciplina}|{nome}"
        apt[chave] = qtde_alunos

        # Mantém na lista global
        lista = self._dados.setdefault("turmas", [])
        if nome not in lista:
            lista.append(nome)
            
        if auto_save: self._salvar()

    def obter_alunos_turma(self, disciplina: str, turma: str) -> int:
        """Retorna a quantidade padrão de alunos para uma turma."""
        chave = f"{disciplina.upper()}|{turma.upper()}"
        return self._dados.get("alunos_por_turma", {}).get(chave, 0)

    def editar_disciplina(self, nome_antigo: str, nome_novo: str, curso: str):
        """Renomeia uma disciplina e atualiza turmas e aulas vinculadas."""
        nome_antigo = nome_antigo.strip().upper()
        nome_novo = nome_novo.strip().upper()
        curso = curso.strip().upper()
        if not nome_novo or nome_antigo == nome_novo:
            return

        # 1. disciplinas_por_curso
        dpc = self._dados.get("disciplinas_por_curso", {})
        if curso in dpc and nome_antigo in dpc[curso]:
            idx = dpc[curso].index(nome_antigo)
            dpc[curso][idx] = nome_novo

        # 2. Lista global de disciplinas
        lista = self._dados.get("disciplinas", [])
        for i, d in enumerate(lista):
            if d.strip().upper() == nome_antigo:
                lista[i] = nome_novo
                break

        # 3. turmas_por_disciplina
        tpd = self._dados.get("turmas_por_disciplina", {})
        if nome_antigo in tpd:
            tpd[nome_novo] = tpd.pop(nome_antigo)

        # 4. alunos_por_turma (chave tem "DISCIPLINA|TURMA")
        apt = self._dados.get("alunos_por_turma", {})
        novas_apt = {}
        for k, v in apt.items():
            partes = k.split("|", 1)
            if partes[0].strip().upper() == nome_antigo:
                novas_apt[f"{nome_novo}|{partes[1]}"] = v
            else:
                novas_apt[k] = v
        self._dados["alunos_por_turma"] = novas_apt

        # 5. Aulas
        for aula in self._dados.get("aulas", []):
            if str(aula.get("disciplina", "")).strip().upper() == nome_antigo:
                aula["disciplina"] = nome_novo

        self._salvar()

    def excluir_disciplina(self, nome: str, curso: str):
        """Remove uma disciplina e suas turmas vinculadas (não remove aulas)."""
        nome = nome.strip().upper()
        curso = curso.strip().upper()

        # 1. disciplinas_por_curso
        dpc = self._dados.get("disciplinas_por_curso", {})
        if curso in dpc and nome in dpc[curso]:
            dpc[curso].remove(nome)

        # 2. Lista global
        self._dados["disciplinas"] = [
            d for d in self._dados.get("disciplinas", [])
            if d.strip().upper() != nome
        ]

        # 3. turmas_por_disciplina
        tpd = self._dados.get("turmas_por_disciplina", {})
        tpd.pop(nome, None)

        # 4. alunos_por_turma
        apt = self._dados.get("alunos_por_turma", {})
        self._dados["alunos_por_turma"] = {
            k: v for k, v in apt.items()
            if k.split("|", 1)[0].strip().upper() != nome
        }

        self._salvar()

    def editar_turma(self, disciplina: str, nome_antigo: str, nome_novo: str, qtde_alunos: int):
        """Renomeia uma turma e atualiza aulas vinculadas."""
        disciplina = disciplina.strip().upper()
        nome_antigo = nome_antigo.strip().upper()
        nome_novo = nome_novo.strip().upper()
        if not nome_novo or nome_antigo == nome_novo:
            # Mesmo nome — apenas atualiza qtde_alunos
            chave = f"{disciplina}|{nome_antigo}"
            self._dados.setdefault("alunos_por_turma", {})[chave] = qtde_alunos
            self._salvar()
            return

        # 1. turmas_por_disciplina
        tpd = self._dados.get("turmas_por_disciplina", {})
        if disciplina in tpd and nome_antigo in tpd[disciplina]:
            idx = tpd[disciplina].index(nome_antigo)
            tpd[disciplina][idx] = nome_novo

        # 2. Lista global
        lista = self._dados.get("turmas", [])
        for i, t in enumerate(lista):
            if t.strip().upper() == nome_antigo:
                lista[i] = nome_novo
                break

        # 3. alunos_por_turma
        apt = self._dados.get("alunos_por_turma", {})
        chave_ant = f"{disciplina}|{nome_antigo}"
        chave_nov = f"{disciplina}|{nome_novo}"
        apt.pop(chave_ant, None)
        apt[chave_nov] = qtde_alunos

        # 4. Aulas
        for aula in self._dados.get("aulas", []):
            disc_aula = str(aula.get("disciplina", "")).strip().upper()
            turm_aula = str(aula.get("turma", "")).strip().upper()
            if disc_aula == disciplina and turm_aula == nome_antigo:
                aula["turma"] = nome_novo

        self._salvar()

    def excluir_turma(self, disciplina: str, nome: str):
        """Remove uma turma (não remove as aulas vinculadas)."""
        disciplina = disciplina.strip().upper()
        nome = nome.strip().upper()

        # 1. turmas_por_disciplina
        tpd = self._dados.get("turmas_por_disciplina", {})
        if disciplina in tpd and nome in tpd[disciplina]:
            tpd[disciplina].remove(nome)

        # 2. Lista global
        self._dados["turmas"] = [
            t for t in self._dados.get("turmas", [])
            if t.strip().upper() != nome
        ]

        # 3. alunos_por_turma
        apt = self._dados.get("alunos_por_turma", {})
        apt.pop(f"{disciplina}|{nome}", None)

        self._salvar()

    def importar_cores(self, rows: List[dict]):
        """Atualiza cores dos cursos a partir de uma lista de dicionários (curso, cor)."""
        cores = self._dados.setdefault("cores_cursos", {})
        count = 0
        for row in rows:
            curso = row.get("curso", "").strip().upper()
            cor = row.get("cor", "").strip()
            if curso and cor:
                cor_f = formatar_cor_hex(cor)
                cores[curso] = cor_f
                # Atualizar em aulas também para consistência imediata
                for aula in self._dados.get("aulas", []):
                    if str(aula.get("curso", "")).strip().upper() == curso:
                        aula["cor_fundo"] = cor_f
                count += 1
        if count > 0:
            self._salvar()
        return count

    # ------------------------------------------------------------------
    # CRUD — Professores
    # ------------------------------------------------------------------

    def listar_professores(self) -> List[str]:
        """Retorna lista de professores únicos das aulas, ordenados."""
        profs = set()
        for aula in self._dados.get("aulas", []):
            p = aula.get("professor")
            if p and str(p).strip():
                profs.add(str(p).strip().upper())
        return sorted(profs)

    def editar_professor(self, nome_antigo: str, nome_novo: str):
        """Renomeia um professor em todas as aulas vinculadas."""
        nome_antigo = nome_antigo.strip().upper()
        nome_novo = nome_novo.strip().upper()
        if not nome_novo:
            return
        for aula in self._dados.get("aulas", []):
            if str(aula.get("professor", "")).strip().upper() == nome_antigo:
                aula["professor"] = nome_novo
        self._salvar()

    # ------------------------------------------------------------------
    # Compatibilidade
    # ------------------------------------------------------------------

    def recarregar(self):
        """Relê o arquivo JSON do disco, atualizando os dados em memória se a leitura for bem-sucedida."""
        novos_dados = self._carregar_arquivo()
        if novos_dados is not None:
            self._dados = novos_dados

    def fechar(self):
        """Sem conexão para fechar no modo JSON. Mantido por compatibilidade."""
        pass

    def importar_dados_csv(self, rows: List[dict], callback=None):
        """
        Processa linhas do CSV e cadastra a hierarquia.
        A cada linha, chama callback(atual, total). Se retornar False, cancela.
        """
        total = len(rows)
        for i, row in enumerate(rows):
            if callback and not callback(i + 1, total):
                break # Cancelado pelo usuário

            faculdade = row.get("Faculdade", "").strip().upper()
            curso = row.get("Curso", "").strip().upper()
            disciplina = row.get("Disciplina", "").strip().upper()
            turma = row.get("Turma", "").strip().upper()
            try:
                qtde = int(row.get("Alunos", "0"))
            except ValueError:
                qtde = 0

            if not faculdade or not curso:
                continue

            # Chama adicionar sem salvar a cada vez (performático)
            self.adicionar_faculdade(faculdade, auto_save=False)
            self.adicionar_curso(faculdade, curso, auto_save=False)
            if disciplina:
                self.adicionar_disciplina(curso, disciplina, auto_save=False)
            if turma and disciplina:
                self.adicionar_turma(disciplina, turma, qtde_alunos=qtde, auto_save=False)
        
        # Salva tudo no final
        self._salvar()

    def exportar_para_excel(self, destination):
        """Exporta todos os dados cadastrados para um arquivo Excel (.xlsx).
        'destination' pode ser uma string (caminho) ou um objeto de arquivo aberto em modo 'wb'.
        """
        wb = Workbook()
        
        # 1. Aba de Aulas
        ws_aulas = wb.active
        ws_aulas.title = "Aulas"
        ws_aulas.append([
            "ID", "Laboratório", "Dia", "Início", "Fim", 
            "Faculdade", "Curso", "Disciplina", "Turma", "Alunos", "Professor"
        ])
        
        for a in self.listar_todas_aulas():
            ws_aulas.append([
                a.id, a.laboratorio, a.dia_semana, a.hora_inicio, a.hora_fim,
                a.faculdade, a.curso, a.disciplina, a.turma, a.qtde_alunos, a.professor
            ])

        # 2. Aba de Cursos
        ws_cursos = wb.create_sheet("Cursos")
        ws_cursos.append(["Faculdade", "Curso", "Cor HEX"])
        for c in self.obter_todos_cursos():
            ws_cursos.append([c["faculdade"], c["nome"], c["cor"]])

        # 3. Aba de Instituições
        ws_facs = wb.create_sheet("Instituições")
        ws_facs.append(["Nome da Faculdade"])
        for f in self.listar_faculdades():
            ws_facs.append([f])

        wb.save(destination)

    def exportar_estatisticas_excel(self, destination):
        """Exporta os dados estatísticos do dashboard para Excel."""
        from openpyxl import Workbook
        from datetime import datetime
        
        stats = self.obter_estatisticas_aulas()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Aulas"
        
        # Cabeçalhos
        ws.append(["Métrica", "Quantidade", "Data de Extração"])
        
        data_extracao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        # Dados
        itens = [
            ("Quantidade de Aulas Semestrais", stats["semestral"]),
            ("Quantidade de Aulas Eventuais", stats["eventual"]),
            ("Aulas de Reposição", stats["reposicao"] if stats["reposicao"] is not None else "Sem dados"),
            ("Aulas Canceladas", stats["cancelada"] if stats["cancelada"] is not None else "Sem dados"),
            ("Aulas Indeferidas", stats["indeferida"] if stats["indeferida"] is not None else "Sem dados")
        ]
        
        for metrica, valor in itens:
            ws.append([metrica, valor, data_extracao])
            
        # Ajuste de largura básico
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        
        wb.save(destination)

    def exportar_para_csv(self, destination):
        """Exporta todos os dados das aulas para um arquivo CSV.
        'destination' pode ser uma string (caminho) ou um objeto de arquivo aberto.
        """
        if isinstance(destination, str):
            f = open(destination, "w", encoding="utf-8-sig", newline="")
            close_it = True
        else:
            f = destination
            close_it = False

        try:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([
                "Laboratório", "Dia", "Início", "Fim", 
                "Faculdade", "Curso", "Disciplina", "Turma", "Alunos", "Professor"
            ])
            for a in self.listar_todas_aulas():
                writer.writerow([
                    a.laboratorio, a.dia_semana, a.hora_inicio, a.hora_fim,
                    a.faculdade, a.curso, a.disciplina, a.turma, a.qtde_alunos, a.professor
                ])
        finally:
            if close_it:
                f.close()
    # ------------------------------------------------------------------
    # Status de Laboratórios (Ligado/Desligado)
    # ------------------------------------------------------------------

    def obter_todos_status_labs(self) -> dict:
        """Retorna o dicionário de status atual de todos os labs."""
        self.recarregar() # Garante que lemos a versão mais recente do disco
        return self._dados.get("status_laboratorios", {})

    def atualizar_status_lab(self, lab_nome: str, status: str):
        """Atualiza o status de um laboratório específico."""
        self.recarregar() # Evita sobrepor alterações de outros usuários
        self._dados.setdefault("status_laboratorios", {})[lab_nome] = status
        self._salvar()
    def obter_status_horario(self) -> str:
        """Retorna o horário da próxima aula (calculado dinamicamente)."""
        return self.calcular_proximo_horario_aula()

    def obter_proximo_horario_detalhado(self) -> dict:
        """Retorna dicionário com 'dia' e 'horario' da próxima aula."""
        from datetime import datetime, timedelta
        now = datetime.now()
        dias_pt = ["SEGUNDA-FEIRA", "TERÇA-FEIRA", "QUARTA-FEIRA", "QUINTA-FEIRA", "SEXTA-FEIRA", "SÁBADO", "DOMINGO"]
        hoje_idx = now.weekday()
        hora_atual = now.strftime("%H:%M")
        hoje_data_str = now.strftime("%Y-%m-%d")
        hoje_data_alt = now.strftime("%d/%m/%Y")
        
        aulas = self.listar_todas_aulas()
        if not aulas:
            return {"dia": dias_pt[now.weekday() % 6], "horario": "07:30"}

        # 1. Hoje
        dia_hoje = dias_pt[hoje_idx]
        aulas_hoje = []
        for a in aulas:
            if a.dia_semana != dia_hoje or a.hora_inicio < hora_atual:
                continue
            
            # Se for eventual, só conta se for hoje
            if a.is_eventual:
                if a.data_eventual not in [hoje_data_str, hoje_data_alt]:
                    continue
            
            aulas_hoje.append(a)

        if aulas_hoje:
            return {"dia": dia_hoje, "horario": min(a.hora_inicio for a in aulas_hoje)}
            
        # 2. Próximos dias
        for i in range(1, 8):
            p_idx = (hoje_idx + i) % 7
            dia_b = dias_pt[p_idx]
            if dia_b == "DOMINGO": continue
            
            # Calcula a data do dia de busca
            data_busca = now + timedelta(days=i)
            db_str = data_busca.strftime("%Y-%m-%d")
            db_alt = data_busca.strftime("%d/%m/%Y")
            
            aulas_dia = []
            for a in aulas:
                if a.dia_semana != dia_b:
                    continue
                
                # Se for eventual, só conta se for a data exata
                if a.is_eventual:
                    if a.data_eventual not in [db_str, db_alt]:
                        continue
                
                aulas_dia.append(a)

            if aulas_dia:
                return {"dia": dia_b, "horario": min(a.hora_inicio for a in aulas_dia)}
                
        return {"dia": dias_pt[0], "horario": "07:30"}

    def calcular_proximo_horario_aula(self) -> str:
        """Calcula o horário de início da próxima aula mais próxima do agora."""
        from datetime import datetime, timedelta
        now = datetime.now()
        
        # Mapeamento do weekday do Python (0=Mon, 6=Sun) para o formato do sistema
        dias_pt = [
            "SEGUNDA-FEIRA", "TERÇA-FEIRA", "QUARTA-FEIRA", 
            "QUINTA-FEIRA", "SEXTA-FEIRA", "SÁBADO", "DOMINGO"
        ]
        
        hoje_idx = now.weekday()
        hora_atual = now.strftime("%H:%M")
        hoje_data_str = now.strftime("%Y-%m-%d")
        hoje_data_alt = now.strftime("%d/%m/%Y")
        
        aulas = self.listar_todas_aulas()
        if not aulas:
            return "07:30"
            
        # 1. Tenta encontrar hoje, após o horário atual
        dia_hoje = dias_pt[hoje_idx]
        aulas_hoje = []
        for a in aulas:
            if a.dia_semana == dia_hoje and a.hora_inicio >= hora_atual:
                # Se for eventual, só considera se for HOJE
                if a.is_eventual:
                    if a.data_eventual in [hoje_data_str, hoje_data_alt]:
                        aulas_hoje.append(a)
                else:
                    aulas_hoje.append(a)
        
        if aulas_hoje:
            return min(a.hora_inicio for a in aulas_hoje)
            
        # 2. Se não encontrou hoje, busca nos próximos dias (máximo 7 dias à frente)
        for i in range(1, 8):
            proximo_idx = (hoje_idx + i) % 7
            dia_busca = dias_pt[proximo_idx]
            
            # Pula Domingo se não estiver no sistema
            if dia_busca == "DOMINGO":
                continue
            
            # Calcula data para conferir 'aulas eventuais'
            data_alvo = now + timedelta(days=i)
            da_str = data_alvo.strftime("%Y-%m-%d")
            da_alt = data_alvo.strftime("%d/%m/%Y")
                
            aulas_dia = []
            for a in aulas:
                if a.dia_semana == dia_busca:
                     if a.is_eventual:
                          # Somente se a data bater
                          if a.data_eventual in [da_str, da_alt]:
                               aulas_dia.append(a)
                     else:
                          aulas_dia.append(a)

            if aulas_dia:
                return min(a.hora_inicio for a in aulas_dia)
                
        return "07:30"

    def atualizar_status_horario(self, valor: str):
        self.recarregar()
        self._dados["status_horario"] = valor
        self._salvar()

    def obter_todas_acoes_labs(self) -> dict:
        self.recarregar()
        return self._dados.get("status_acoes_labs", {})

    def atualizar_acao_lab(self, lab_nome: str, acao: str):
        self.recarregar()
        self._dados.setdefault("status_acoes_labs", {})[lab_nome] = acao
        self._salvar()
