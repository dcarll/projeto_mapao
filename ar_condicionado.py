"""
micros.py  —  Planilha "Micros"  (arquivo autossuficiente)

Execute:  python micros.py
Requer:   Python 3.8+  (tkinter já incluso no Python padrão)
Opcional: pip install openpyxl   → habilita abrir/exportar .xlsx
"""

# ── diagnóstico de ambiente ──────────────────────────────────────────────────
import sys, os

if sys.version_info < (3, 8):
    print("ERRO: Python 3.8 ou superior é necessário.")
    sys.exit(1)

try:
    import tkinter as tk
    from tkinter import ttk, colorchooser, messagebox, filedialog
except ImportError:
    print("ERRO: tkinter não encontrado.")
    print("  Windows/Mac: reinstale o Python marcando a opção 'tcl/tk'.")
    print("  Linux: sudo apt install python3-tk")
    sys.exit(1)

import json, re, csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── constantes ───────────────────────────────────────────────────────────────
ROWS            = 50
COLS            = 26
CELL_W          = 80
DEFAULT_ROW_H   = 22
HEADER_W        = 45
COL_HEADER_H    = 22
FONT_NAME       = "Calibri"
DEFAULT_FONT_SZ = 11
RESIZE_ZONE     = 4
APP_TITLE       = "Ar-Condicionado"
SAVE_DIR        = "inventarios"
DEFAULT_SAVE    = os.path.join(SAVE_DIR, "ar_condicionado.json")
os.makedirs(SAVE_DIR, exist_ok=True)

COL_LETTERS = [
    chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
    for i in range(COLS)
]

def col_letter(i):
    return COL_LETTERS[i] if 0 <= i < len(COL_LETTERS) else ""

def col_index(s):
    s = s.upper(); v = 0
    for ch in s: v = v * 26 + (ord(ch) - 64)
    return v - 1

def parse_ref(ref):
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", ref.strip())
    if not m: return None
    return (int(m.group(2)) - 1, col_index(m.group(1)))

# ── motor de fórmulas ────────────────────────────────────────────────────────
class FormulaEngine:
    def __init__(self, model): self.m = model

    def evaluate(self, expr):
        if not expr.startswith("="): return expr
        e = expr[1:].strip().upper()
        for pat, fn in [
            (r"(?:SOMA|SUM)\((.+)\)",               self._sum),
            (r"(?:M[EÉ]DIA|AVERAGE)\((.+)\)",       self._avg),
            (r"(?:M[AÁ]X(?:IMO)?|MAX)\((.+)\)",     self._max),
            (r"(?:M[IÍ]N(?:IMO)?|MIN)\((.+)\)",     self._min),
            (r"(?:CONT\.?N[UÚ]M|COUNT)\((.+)\)",    self._count),
        ]:
            mm = re.fullmatch(pat, e)
            if mm: return fn(mm.group(1))
        return "#FUNC?"

    def _vals(self, arg):
        out = []
        for part in re.split(r"[;,]", arg):
            part = part.strip()
            rm = re.fullmatch(r"([A-Z]+\d+):([A-Z]+\d+)", part)
            if rm:
                r1,c1 = parse_ref(rm.group(1)) or (0,0)
                r2,c2 = parse_ref(rm.group(2)) or (0,0)
                for r in range(min(r1,r2), max(r1,r2)+1):
                    for c in range(min(c1,c2), max(c1,c2)+1):
                        n = self._n(self.m.get(r,c))
                        if n is not None: out.append(n)
            else:
                ref = parse_ref(part)
                n = self._n(self.m.get(*ref) if ref else part)
                if n is not None: out.append(n)
        return out

    @staticmethod
    def _n(s):
        try: return float(str(s).replace(",",".").strip())
        except: return None

    @staticmethod
    def _fmt(v):
        return str(int(v)) if v == int(v) else f"{v:.6f}".rstrip("0").rstrip(".")

    def _sum(self, a):
        v = self._vals(a); return self._fmt(sum(v)) if v else "0"
    def _avg(self, a):
        v = self._vals(a); return self._fmt(sum(v)/len(v)) if v else "#DIV/0!"
    def _max(self, a):
        v = self._vals(a); return self._fmt(max(v)) if v else ""
    def _min(self, a):
        v = self._vals(a); return self._fmt(min(v)) if v else ""
    def _count(self, a):
        return str(len(self._vals(a)))

# ── modelo ───────────────────────────────────────────────────────────────────
class SheetModel:
    def __init__(self):
        self.data        = {}   # (r,c) → str
        self.styles      = {}   # (r,c) → dict
        self.merges      = {}   # (r,c) master → (r2,c2)
        self.row_heights = {}   # r → px
        self.col_widths  = {}   # c → px
        self._eng        = FormulaEngine(self)
        self.undo_stack  = []
        self.redo_stack  = []

    def save_history(self):
        snap = self.to_dict()
        self.undo_stack.append(snap)
        if len(self.undo_stack) > 100: self.undo_stack.pop(0)
        self.redo_stack.clear()

    def undo(self):
        if not self.undo_stack: return False
        curr = self.to_dict()
        snap = self.undo_stack.pop()
        self.redo_stack.append(curr)
        self.from_dict(snap)
        return True

    def redo(self):
        if not self.redo_stack: return False
        curr = self.to_dict()
        snap = self.redo_stack.pop()
        self.undo_stack.append(curr)
        self.from_dict(snap)
        return True

    def get(self, r, c):       return self.data.get((r,c), "")
    def get_disp(self, r, c):
        raw = self.data.get((r,c), "")
        return self._eng.evaluate(raw) if raw.startswith("=") else raw
    def set(self, r, c, v):    self.data[(r,c)] = v
    def get_style(self, r, c): return self.styles.get((r,c), {}).copy()
    def set_style(self, r, c, s): self.styles[(r,c)] = s

    def default_style(self):
        return {"bold":False,"italic":False,"font_size":DEFAULT_FONT_SZ,
                "bg":"#ffffff","fg":"#000000","align":"w",
                "bt":False, "bb":False, "bl":False, "br":False}

    def merged_style(self, r, c):
        d = self.default_style(); d.update(self.styles.get((r,c),{})); return d

    def row_h(self, r):   return self.row_heights.get(r, DEFAULT_ROW_H)
    def row_y(self, r):   return COL_HEADER_H + sum(self.row_h(i) for i in range(r))
    def col_w(self, c):   return self.col_widths.get(c, CELL_W)
    def col_x(self, c):   return HEADER_W + sum(self.col_w(i) for i in range(c))

    def is_master(self, r, c):  return (r,c) in self.merges
    def master_of(self, r, c):
        for (mr,mc),(r2,c2) in self.merges.items():
            if mr<=r<=r2 and mc<=c<=c2 and (r,c)!=(mr,mc): return (mr,mc)
        return None

    def merge_range(self, r1, c1, r2, c2):
        rn,rx = min(r1,r2),max(r1,r2); cn,cx = min(c1,c2),max(c1,c2)
        for k in [k for k in self.merges if rn<=k[0]<=rx and cn<=k[1]<=cx]:
            del self.merges[k]
        if rn==rx and cn==cx: return
        self.merges[(rn,cn)] = (rx,cx)
        for r in range(rn,rx+1):
            for c in range(cn,cx+1):
                if (r,c)!=(rn,cn):
                    self.data.pop((r,c),None); self.styles.pop((r,c),None)

    def unmerge_range(self, r1, c1, r2, c2):
        rn,rx = min(r1,r2),max(r1,r2); cn,cx = min(c1,c2),max(c1,c2)
        for k in [k for k in self.merges if rn<=k[0]<=rx and cn<=k[1]<=cx]:
            del self.merges[k]

    def to_dict(self):
        return {
            "data":        {f"{r},{c}":v for (r,c),v in self.data.items()},
            "styles":      {f"{r},{c}":s for (r,c),s in self.styles.items()},
            "merges":      {f"{r},{c}":list(v) for (r,c),v in self.merges.items()},
            "row_heights": {str(r):h for r,h in self.row_heights.items()},
            "col_widths":  {str(c):w for c,w in self.col_widths.items()},
        }

    def from_dict(self, d):
        self.data={}; self.styles={}; self.merges={}; self.row_heights={}; self.col_widths={}
        for k,v in d.get("data",{}).items():
            r,c=map(int,k.split(",")); self.data[(r,c)]=v
        for k,s in d.get("styles",{}).items():
            r,c=map(int,k.split(",")); self.styles[(r,c)]=s
        for k,v in d.get("merges",{}).items():
            r,c=map(int,k.split(",")); self.merges[(r,c)]=tuple(v)
        for k,v in d.get("row_heights",{}).items():
            self.row_heights[int(k)]=int(v)
        for k,v in d.get("col_widths",{}).items():
            self.col_widths[int(k)]=int(v)

    def reset(self):
        self.__init__()

# ── gerenciador de arquivos ──────────────────────────────────────────────────
class FileManager:
    def __init__(self, widget): self.w = widget

    # ── salvar / abrir JSON ──────────────────
    def save(self):
        if not self.w.save_path: self.save_as(); return
        self._write(self.w.save_path)

    def save_as(self):
        p = filedialog.asksaveasfilename(
            title="Salvar como", defaultextension=".json",
            filetypes=[("Planilha JSON","*.json"),("Todos","*.*")])
        if p: self.w.save_path = p; self._write(p)

    def _write(self, path):
        with open(path,"w",encoding="utf-8") as f:
            json.dump(self.w.model.to_dict(), f, ensure_ascii=False, indent=2)
        self.w.winfo_toplevel().title(f"{APP_TITLE} — {os.path.basename(path)}")

    def load_json(self, path=None):
        p = path or self.w.save_path
        if p and os.path.exists(p):
            try:
                with open(p,"r",encoding="utf-8") as f:
                    self.w.model.from_dict(json.load(f))
                self.w.render_all()
                self.w.winfo_toplevel().title(f"{APP_TITLE} — {os.path.basename(p)}")
            except Exception as e:
                messagebox.showerror("Erro ao carregar", str(e))

    # ── abrir (detecta tipo) ─────────────────
    def open_file(self):
        p = filedialog.askopenfilename(
            title="Abrir",
            filetypes=[("Todos suportados","*.json *.xlsx *.xls *.csv *.tsv"),
                       ("JSON","*.json"),("Excel","*.xlsx *.xls"),
                       ("CSV","*.csv *.tsv"),("Todos","*.*")])
        if not p: return
        ext = os.path.splitext(p)[1].lower()
        if ext in (".xlsx",".xls"):   self._load_xlsx(p)
        elif ext in (".csv",".tsv"):  self._load_csv(p)
        else:
            self.w.save_path = p; self.load_json(p)

    # ── importar CSV ─────────────────────────
    def import_csv(self):
        p = filedialog.askopenfilename(
            title="Importar CSV",
            filetypes=[("CSV / TSV","*.csv *.tsv"),("Todos","*.*")])
        if p: self._load_csv(p)

    def _load_csv(self, path):
        self.w.model.reset()
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                sep = "\t" if path.lower().endswith(".tsv") else ","
                for r, row in enumerate(csv.reader(f, delimiter=sep)):
                    for c, val in enumerate(row):
                        if val: self.w.model.set(r, c, val)
            self.w.render_all()
        except Exception as e:
            messagebox.showerror("Erro ao importar CSV", str(e))

    # ── exportar CSV ─────────────────────────
    def export_csv(self):
        p = filedialog.asksaveasfilename(
            title="Exportar CSV", defaultextension=".csv",
            filetypes=[("CSV","*.csv"),("Todos","*.*")])
        if not p: return
        try:
            mr = max((r for r,c in self.w.model.data), default=0)+1
            mc = max((c for r,c in self.w.model.data), default=0)+1
            with open(p,"w",newline="",encoding="utf-8-sig") as f:
                w = csv.writer(f)
                for r in range(mr):
                    w.writerow([self.w.model.get_disp(r,c) for c in range(mc)])
            messagebox.showinfo("Exportado", f"CSV salvo em:\n{p}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar CSV", str(e))

    # ── abrir xlsx ───────────────────────────
    def _load_xlsx(self, path=None):
        if not HAS_OPENPYXL:
            messagebox.showwarning("openpyxl não instalado",
                "Para usar arquivos .xlsx instale:\n\n  pip install openpyxl"); return
        p = path or filedialog.askopenfilename(
            title="Abrir Excel",
            filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not p: return
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            ws = wb.active; self.w.model.reset()
            for ri, row in enumerate(ws.iter_rows()):
                for ci, cell in enumerate(row):
                    if cell.value is not None:
                        self.w.model.set(ri, ci, str(cell.value))
                    st = {}
                    if cell.font:
                        if cell.font.bold:      st["bold"]      = True
                        if cell.font.italic:    st["italic"]    = True
                        if cell.font.size:      st["font_size"] = int(cell.font.size)
                        if (cell.font.color and cell.font.color.type == "rgb"
                                and len(cell.font.color.rgb) == 8):
                            st["fg"] = "#" + cell.font.color.rgb[2:]
                    if (cell.fill and cell.fill.fgColor
                            and cell.fill.fgColor.type == "rgb"
                            and len(cell.fill.fgColor.rgb) == 8
                            and cell.fill.fgColor.rgb[2:].upper() != "000000"):
                        st["bg"] = "#" + cell.fill.fgColor.rgb[2:]
                    if st: self.w.model.set_style(ri, ci, st)
            self.w.render_all()
        except Exception as e:
            messagebox.showerror("Erro ao abrir xlsx", str(e))

    # ── exportar xlsx ────────────────────────
    def export_xlsx(self):
        if not HAS_OPENPYXL:
            messagebox.showwarning("openpyxl não instalado",
                "Para exportar .xlsx instale:\n\n  pip install openpyxl"); return
        p = filedialog.asksaveasfilename(
            title="Exportar Excel", defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("Todos","*.*")])
        if not p: return
        try:
            wb = openpyxl.Workbook(); ws = wb.active
            mr = max((r for r,c in self.w.model.data), default=0)+1
            mc = max((c for r,c in self.w.model.data), default=0)+1
            am = {"w":"left","center":"center","e":"right"}
            for r in range(mr):
                for c in range(mc):
                    v = self.w.model.get_disp(r,c)
                    if not v: continue
                    cell = ws.cell(row=r+1, column=c+1, value=v)
                    st   = self.w.model.merged_style(r,c)
                    cell.font      = Font(bold=st["bold"], italic=st["italic"],
                                         size=st["font_size"],
                                         color=st["fg"].lstrip("#").upper())
                    bg = st["bg"].lstrip("#").upper()
                    if bg != "FFFFFF":
                        cell.fill  = PatternFill("solid", fgColor=bg)
                    cell.alignment = Alignment(horizontal=am.get(st["align"],"left"))
            for (r,c),(r2,c2) in self.w.model.merges.items():
                ws.merge_cells(start_row=r+1,start_column=c+1,
                               end_row=r2+1, end_column=c2+1)

            # Ajusta larguras e alturas no Excel
            for c in range(mc):
                let = col_letter(c)
                if let: ws.column_dimensions[let].width = self.w.model.col_w(c) / 7.0
            for r in range(mr):
                ws.row_dimensions[r+1].height = self.w.model.row_h(r) * 0.75

            wb.save(p)
            messagebox.showinfo("Exportado", f"Excel salvo em:\n{p}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar xlsx", str(e))

# ── widget principal ─────────────────────────────────────────────────────────
class SpreadsheetWidget(tk.Frame):
    def __init__(self, parent, model, save_path, **kw):
        self.use_menubar = kw.pop("use_menubar", True)
        super().__init__(parent, **kw)
        self.model     = model
        self.save_path = save_path
        self.rows      = ROWS
        self.cols      = COLS
        self.file_mgr  = FileManager(self)

        self.sel_start = None
        self.sel_end   = None
        self.editing   = None
        self.clipboard = None
        self.cut_cells = None
        self._resize_idx   = None
        self._resize_type  = None
        self._resize_start = 0
        self._resize_d0    = 0
        self._is_resizing  = False

        if self.use_menubar:
            self._build_menubar()
        self._build_toolbar()
        self._build_grid()
        self._bind_keys()
        self.render_all()

    # ── menu arquivo ─────────────────────────
    def _build_menubar(self):
        root = self.winfo_toplevel()
        mb   = tk.Menu(root)
        root.configure(menu=mb)

        fm = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="Arquivo", menu=fm)
        fm.add_command(label="Abrir…",                    command=self.file_mgr.open_file)
        fm.add_command(label="Importar CSV…",             command=self.file_mgr.import_csv)
        fm.add_command(label="Exportar CSV…",             command=self.file_mgr.export_csv)
        fm.add_separator()
        xlsx_sfx = "" if HAS_OPENPYXL else "  [pip install openpyxl]"
        fm.add_command(label=f"Abrir Excel (.xlsx)…{xlsx_sfx}",
                       command=self.file_mgr._load_xlsx)
        fm.add_command(label=f"Exportar Excel (.xlsx)…{xlsx_sfx}",
                       command=self.file_mgr.export_xlsx)
        fm.add_separator()
        fm.add_command(label="Salvar             Ctrl+S", command=self.file_mgr.save)
        fm.add_command(label="Salvar como…",              command=self.file_mgr.save_as)
        fm.add_separator()
        fm.add_command(label="Sair", command=root.destroy)

    # ── toolbar ──────────────────────────────
    def _build_toolbar(self):
        tb = tk.Frame(self, bg="#f0f0f0", bd=1, relief="raised")
        tb.pack(fill="x", side="top")
        bs = {"relief":"flat","padx":5,"pady":2,"bg":"#f0f0f0","activebackground":"#d0d0d0"}

        def btn(text, cmd, font=None):
            b = tk.Button(tb, text=text, command=cmd,
                          font=font or (FONT_NAME,10), **bs)
            b.pack(side="left", padx=2, pady=2); return b

        def sep():
            tk.Label(tb,text="|",bg="#f0f0f0",fg="#bbb").pack(side="left",padx=1)

        btn("N",           self.toggle_bold,          font=(FONT_NAME,10,"bold"))
        btn("I",           self.toggle_italic,         font=(FONT_NAME,10,"italic"))
        sep()
        btn("Fundo",        self.pick_bg)
        btn("Cor",       self.pick_fg)
        sep()
        btn("◀",           self.align_left)
        btn("▬",           self.align_center)
        btn("▶",           self.align_right)
        sep()
        btn("A+",          self.font_increase)
        btn("A-",          self.font_decrease)
        sep()
        btn("⊞ Mesclar",   self.merge_selection)
        btn("⊟ Desmesclar",self.unmerge_selection)
        sep()
        btn("田 Todas",    self.toggle_all_borders)
        btn("囗 Contorno", self.toggle_borders)
        sep()
        btn("Σ SOMA",      self._insert_sum)
        sep()
        btn("📥 Abrir",    self.file_mgr.open_file)
        btn("📤 CSV",      self.file_mgr.export_csv)
        btn("📊 Exportar",     self.file_mgr.export_xlsx)
        sep()
        btn("💾",          self.file_mgr.save)

        # barra de fórmula
        ff = tk.Frame(self, bg="#f0f0f0")
        ff.pack(fill="x", side="top")
        self.cell_ref_var = tk.StringVar(value="A1")
        tk.Label(ff, textvariable=self.cell_ref_var, width=7,
                 font=(FONT_NAME,10), bg="#dde", relief="sunken"
                 ).pack(side="left", padx=4, pady=1)
        tk.Label(ff, text="fx", bg="#f0f0f0",
                 font=(FONT_NAME,10,"italic"), fg="#555"
                 ).pack(side="left", padx=2)
        self.formula_var = tk.StringVar()
        fe = tk.Entry(ff, textvariable=self.formula_var,
                      font=(FONT_NAME,10), relief="sunken", bd=1)
        fe.pack(side="left", fill="x", expand=True, padx=2, pady=1)
        fe.bind("<Return>", self._formula_commit)

    def _insert_sum(self):
        if not self.sel_start: return
        if self.sel_end and self.sel_end != self.sel_start:
            r1,c1 = self.sel_start; r2,c2 = self.sel_end
            ref = (f"{col_letter(min(c1,c2))}{min(r1,r2)+1}:"
                   f"{col_letter(max(c1,c2))}{max(r1,r2)+1}")
        else:
            r,c = self.sel_start
            ref = f"{col_letter(c)}{r+1}:{col_letter(c)}{r+1}"
        self.formula_var.set(f"=SOMA({ref})")

    # ── grid / canvas ────────────────────────
    def _build_grid(self):
        cont = tk.Frame(self); cont.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(cont, bg="#ffffff", cursor="arrow",
                                highlightthickness=0)
        vb = ttk.Scrollbar(cont, orient="vertical",   command=self.canvas.yview)
        hb = ttk.Scrollbar(cont, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vb.set, xscrollcommand=hb.set)
        hb.pack(side="bottom", fill="x")
        vb.pack(side="right",  fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self._upd_scroll()

        self.entry_var    = tk.StringVar()
        self.inline_entry = tk.Entry(self.canvas, textvariable=self.entry_var,
                                     font=(FONT_NAME,DEFAULT_FONT_SZ),
                                     relief="flat", bd=0,
                                     highlightthickness=1,
                                     highlightbackground="#1565c0")
        self.inline_entry.bind("<Return>", self._entry_commit)
        self.inline_entry.bind("<Escape>", self._entry_cancel)
        self.inline_entry.bind("<Tab>",    self._entry_tab)

        self.canvas.bind("<ButtonPress-1>",   self._on_click)
        self.canvas.bind("<B1-Motion>",       self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<Double-Button-1>", self._on_dblclick)
        self.canvas.bind("<Button-3>",        self._on_rclick)
        self.canvas.bind("<Motion>",          self._on_motion)
        self.canvas.bind("<Key>",             self._on_key)
        self.canvas.bind("<MouseWheel>",      self._on_wheel)
        self.canvas.bind("<Button-4>",        self._on_wheel)
        self.canvas.bind("<Button-5>",        self._on_wheel)

    def _bind_keys(self):
        for seq, fn in [
            ("<Control-c>",self.copy),  ("<Control-C>",self.copy),
            ("<Control-v>",self.paste), ("<Control-V>",self.paste),
            ("<Control-x>",self.cut),   ("<Control-X>",self.cut),
            ("<Control-s>",self.file_mgr.save),
            ("<Control-S>",self.file_mgr.save),
            ("<Control-z>",self.undo),  ("<Control-Z>",self.undo),
            ("<Control-y>",self.redo),  ("<Control-Y>",self.redo),
            ("<Up>",    lambda: self._move_selection(-1, 0)),
            ("<Down>",  lambda: self._move_selection(1, 0)),
            ("<Left>",  lambda: self._move_selection(0, -1)),
            ("<Tab>",    lambda: self._on_tab()),
            ("<Shift-Tab>", lambda: self._on_shifttab()),
            ("<ISO_Left_Tab>", lambda: self._on_shifttab()),
            ("<Delete>", self.delete_selection),
            ("<BackSpace>", self.delete_selection),
            ("<Return>", self._on_enter),
        ]:
            self.canvas.bind(seq, lambda e, f=fn: f())
        self.canvas.focus_set()

    def undo(self):
        if self.model.undo():
            self.render_all()

    def redo(self):
        if self.model.redo():
            self.render_all()

    def _on_tab(self):
        self._move_selection(0, 1)
        return "break"

    def _on_shifttab(self):
        self._move_selection(0, -1)
        return "break"

    def delete_selection(self):
        self.model.save_history()
        for r, c in self._sel_set():
            if self.model.master_of(r, c): continue
            self.model.set(r, c, "")
        self.render_all()

    def _upd_scroll(self):
        tw = self.model.col_x(self.cols) + 4
        th = self.model.row_y(self.rows) + 4
        self.canvas.configure(scrollregion=(0,0,tw,th))

    # ── coordenadas ──────────────────────────
    def _bbox(self, r, c):
        x1 = self.model.col_x(c)
        y1 = self.model.row_y(r)
        if self.model.is_master(r,c):
            r2,c2 = self.model.merges[(r,c)]
            return x1, y1, self.model.col_x(c2+1), self.model.row_y(r2+1)
        return x1, y1, x1+self.model.col_w(c), y1+self.model.row_h(r)

    def _xy_cell(self, cx, cy):
        # Col Header
        if cy < COL_HEADER_H and cx >= HEADER_W:
            acc = HEADER_W
            for c in range(self.cols):
                w = self.model.col_w(c)
                if acc <= cx < acc+w: return ("col_header", c)
                acc += w
        # Row Header
        if cx < HEADER_W and cy >= COL_HEADER_H:
            acc = COL_HEADER_H
            for r in range(self.rows):
                h = self.model.row_h(r)
                if acc <= cy < acc+h: return ("row_header", r)
                acc += h
        # Cells
        if cx >= HEADER_W and cy >= COL_HEADER_H:
            acc_x = HEADER_W; target_c = -1
            for c in range(self.cols):
                w = self.model.col_w(c)
                if acc_x <= cx < acc_x+w:
                    target_c = c; break
                acc_x += w
            if target_c == -1: return None
            
            acc_y = COL_HEADER_H
            for r in range(self.rows):
                h = self.model.row_h(r)
                if acc_y <= cy < acc_y+h:
                    m = self.model.master_of(r, target_c)
                    return m if m else (r, target_c)
                acc_y += h
        return None

    def _cxy(self, e): return self.canvas.canvasx(e.x), self.canvas.canvasy(e.y)

    def _resize_hit(self, cx, cy):
        # Row resize
        if cx < HEADER_W:
            acc = COL_HEADER_H
            for r in range(self.rows):
                acc += self.model.row_h(r)
                if abs(cy-acc) <= RESIZE_ZONE: return ("row", r)
        # Col resize
        if cy < COL_HEADER_H:
            acc = HEADER_W
            for c in range(self.cols):
                acc += self.model.col_w(c)
                if abs(cx-acc) <= RESIZE_ZONE: return ("col", c)
        return None

    # ── render ───────────────────────────────
    def render_all(self):
        self.canvas.delete("all")
        sel = self._sel_set()
        self._upd_scroll()
        tw = self.model.col_x(self.cols)
        th = self.model.row_y(self.rows)
        self.canvas.create_rectangle(0,0,tw,th, fill="#f8f9fa", outline="")

        # cabeçalho colunas
        self.canvas.create_rectangle(0,0,HEADER_W,COL_HEADER_H,
                                     fill="#e8eaed", outline="#cccccc")
        for c in range(self.cols):
            x1 = self.model.col_x(c); x2 = x1+self.model.col_w(c)
            bg = "#b0c4de" if any((r,c) in sel for r in range(self.rows)) else "#e8eaed"
            self.canvas.create_rectangle(x1,0,x2,COL_HEADER_H, fill=bg, outline="#cccccc")
            self.canvas.create_text(x1+(x2-x1)//2, COL_HEADER_H//2,
                                    text=col_letter(c), font=(FONT_NAME,9,"bold"), fill="#333")

        # linhas
        acc_y = COL_HEADER_H
        for r in range(self.rows):
            h = self.model.row_h(r); y1=acc_y; y2=y1+h
            rbg = "#b0c4de" if any((r,c) in sel for c in range(self.cols)) else "#e8eaed"
            self.canvas.create_rectangle(0,y1,HEADER_W,y2, fill=rbg, outline="#cccccc")
            self.canvas.create_text(HEADER_W//2, y1+h//2,
                                    text=str(r+1), font=(FONT_NAME,9), fill="#333")
            for c in range(self.cols): self._draw_cell(r,c,sel)
            acc_y += h

        if self.sel_start:
            r,c = self.sel_start
            self.cell_ref_var.set(f"{col_letter(c)}{r+1}")
            self.formula_var.set(self.model.get(r,c))

    def _draw_cell(self, r, c, sel):
        if self.model.master_of(r,c): return
        x1,y1,x2,y2 = self._bbox(r,c)
        st = self.model.merged_style(r,c)
        if self.model.is_master(r,c):
            r2,c2 = self.model.merges[(r,c)]
            is_sel = any((rr,cc) in sel for rr in range(r,r2+1) for cc in range(c,c2+1))
        else:
            is_sel = (r,c) in sel
        self.canvas.create_rectangle(x1,y1,x2,y2,
                                     fill="#cce5ff" if is_sel else st["bg"],
                                     outline="#d0d0d0")
        
        # Desenha bordas personalizadas (Outline)
        if st.get("bt"): self.canvas.create_line(x1,y1,x2,y1, width=2, fill="#333", tags="border")
        if st.get("bb"): self.canvas.create_line(x1,y2,x2,y2, width=2, fill="#333", tags="border")
        if st.get("bl"): self.canvas.create_line(x1,y1,x1,y2, width=2, fill="#333", tags="border")
        if st.get("br"): self.canvas.create_line(x2,y1,x2,y2, width=2, fill="#333", tags="border")
        txt = self.model.get_disp(r,c)
        if txt:
            fw = "bold"   if st["bold"]   else "normal"
            fi = "italic" if st["italic"] else "roman"
            fn = (FONT_NAME,st["font_size"],fw,fi) if (st["bold"] or st["italic"]) \
                 else (FONT_NAME,st["font_size"])
            a = st.get("align","w")
            tx = x1+3 if a=="w" else (x1+x2)//2 if a=="center" else x2-3
            self.canvas.create_text(tx,(y1+y2)//2, text=txt, font=fn,
                                    fill=st["fg"], anchor=a, width=(x2-x1)-6)
        if self.sel_start==(r,c):
            self.canvas.create_rectangle(x1,y1,x2,y2, outline="#1565c0", width=2)
        if self.model.is_master(r,c):
            self.canvas.create_polygon(x2-8,y1,x2,y1,x2,y1+8, fill="#1565c0",outline="")

    def _sel_set(self):
        if not self.sel_start: return set()
        r1,c1 = self.sel_start; r2,c2 = self.sel_end or self.sel_start
        return {(r,c) for r in range(min(r1,r2),max(r1,r2)+1)
                      for c in range(min(c1,c2),max(c1,c2)+1)}

    # ── eventos mouse ────────────────────────
    def _on_motion(self, e):
        cx,cy = self._cxy(e); hit = self._resize_hit(cx,cy)
        if hit:
            self.canvas.configure(cursor="sb_v_double_arrow" if hit[0]=="row" else "sb_h_double_arrow")
        elif cx < HEADER_W: self.canvas.configure(cursor="arrow")
        else:               self.canvas.configure(cursor="crosshair")

    def _on_click(self, e):
        self.canvas.focus_set(); cx,cy = self._cxy(e)
        hit_resize = self._resize_hit(cx,cy)
        if hit_resize:
            self._is_resizing=True; self._resize_type=hit_resize[0]; self._resize_idx=hit_resize[1]
            self._resize_start=cx if hit_resize[0]=="col" else cy
            self._resize_d0=self.model.col_w(hit_resize[1]) if hit_resize[0]=="col" else self.model.row_h(hit_resize[1])
            return
        self._commit_edit()
        hit = self._xy_cell(cx,cy)
        if hit is None: return
        if isinstance(hit,tuple) and hit[0]=="row_header":
            r=hit[1]; self.sel_start=(r,0); self.sel_end=(r,self.cols-1)
        elif isinstance(hit,tuple) and hit[0]=="col_header":
            c=hit[1]; self.sel_start=(0,c); self.sel_end=(self.rows-1,c)
        else:
            self.sel_start=hit; self.sel_end=hit
        self.render_all()

    def _move_selection(self, dr, dc):
        if not self.sel_start: return
        r, c = self.sel_start
        nr, nc = max(0, min(self.rows-1, r+dr)), max(0, min(self.cols-1, c+dc))
        self.sel_start = (nr, nc); self.sel_end = (nr, nc)
        self.canvas.focus_set()
        self.render_all()
        self._ensure_visible(nr, nc)

    def _ensure_visible(self, r, c):
        # Evita erros se o canvas ainda não estiver mapeado
        if not self.canvas.winfo_width() > 1: return
        
        x1, y1, x2, y2 = self._bbox(r, c)
        cx1 = self.canvas.canvasx(0)
        cx2 = self.canvas.canvasx(self.canvas.winfo_width())
        cy1 = self.canvas.canvasy(0)
        cy2 = self.canvas.canvasy(self.canvas.winfo_height())
        
        full_w = self.cols * CELL_W + HEADER_W
        full_h = self.canvas.bbox("all")[3] if self.canvas.bbox("all") else 1
        
        if x1 < cx1 + HEADER_W:
            self.canvas.xview_moveto((x1 - HEADER_W) / full_w)
        elif x2 > cx2:
            self.canvas.xview_moveto((x2 - self.canvas.winfo_width()) / full_w)
            
        if y1 < cy1 + COL_HEADER_H:
            self.canvas.yview_moveto((y1 - COL_HEADER_H) / full_h)
        elif y2 > cy2:
            self.canvas.yview_moveto((y2 - self.canvas.winfo_height()) / full_h)

    def _on_enter(self):
        if not self.sel_start: return
        r, c = self.sel_start
        self._start_edit(r, c)

    def _on_key(self, e):
        # Ignora teclas de função e navegação
        if e.keysym in ("Up", "Down", "Left", "Right", "Return", "Escape", "Tab", "BackSpace", "Delete", "ISO_Left_Tab"):
            return
        # Inicia edição ao digitar se houver uma célula selecionada
        if not self.sel_start: return
        # Apenas caracteres visíveis, ignorando atalhos Ctrl+...
        if e.char and e.char.isprintable() and not (e.state & 0x4):
            r, c = self.sel_start
            self._start_edit(r, c, clear=True)
            self.entry_var.set(e.char)
            self.inline_entry.icursor("end")

    def _on_drag(self, e):
        cx,cy = self._cxy(e)
        if self._is_resizing:
            delta = (cx if self._resize_type=="col" else cy) - self._resize_start
            new_val = max(8, self._resize_d0 + int(delta))
            if self._resize_type=="col": self.model.col_widths[self._resize_idx] = new_val
            else:                        self.model.row_heights[self._resize_idx] = new_val
            self.render_all(); return
        hit = self._xy_cell(cx,cy)
        if isinstance(hit,tuple) and len(hit)==2 and hit[0] not in ("row_header","col_header"):
            self.sel_end=hit; self.render_all()

    def _on_release(self, e):
        if self._is_resizing:
            self.model.save_history()
        self._is_resizing=False; self._resize_idx=None; self._resize_type=None

    def _on_dblclick(self, e):
        cx,cy = self._cxy(e)
        if self._resize_hit(cx,cy) is not None: return
        hit = self._xy_cell(cx,cy)
        if isinstance(hit,tuple) and len(hit)==2 and hit[0] not in ("row_header","col_header"):
            self.sel_start=hit; self.sel_end=hit; self._start_edit(*hit)

    def _on_rclick(self, e):
        cx,cy = self._cxy(e); hit = self._xy_cell(cx,cy)
        if hit is None: return
        m = tk.Menu(self, tearoff=0)
        if isinstance(hit,tuple) and hit[0]=="row_header":
            r=hit[1]
            m.add_command(label="Inserir linha acima",  command=lambda:self.insert_row(r))
            m.add_command(label="Inserir linha abaixo", command=lambda:self.insert_row(r+1))
            m.add_command(label="Excluir linha",        command=lambda:self.delete_row(r))
            m.add_separator()
            m.add_command(label="Resetar altura",       command=lambda:self._rst_row(r))
        elif isinstance(hit,tuple) and hit[0]=="col_header":
            c=hit[1]
            m.add_command(label="Inserir coluna à esquerda", command=lambda:self.insert_col(c))
            m.add_command(label="Inserir coluna à direita",  command=lambda:self.insert_col(c+1))
            m.add_command(label="Excluir coluna",            command=lambda:self.delete_col(c))
        else:
            m.add_command(label="Copiar  Ctrl+C",  command=self.copy)
            m.add_command(label="Recortar  Ctrl+X",command=self.cut)
            m.add_command(label="Colar  Ctrl+V",   command=self.paste)
            m.add_separator()
            m.add_command(label="⊞ Mesclar",    command=self.merge_selection)
            m.add_command(label="⊟ Desmesclar", command=self.unmerge_selection)
            m.add_separator()
            m.add_command(label="Σ Inserir SOMA", command=self._insert_sum)
        m.tk_popup(e.x_root, e.y_root)

    def _on_wheel(self, e):
        if e.num==4:    self.canvas.yview_scroll(-1,"units")
        elif e.num==5:  self.canvas.yview_scroll(1,"units")
        else:           self.canvas.yview_scroll(-1*(e.delta//120),"units")

    def _rst_row(self, r):
        self.model.row_heights.pop(r,None); self.render_all()

    # ── edição inline ────────────────────────
    def _start_edit(self, r, c, clear=False):
        self.editing=(r,c); x1,y1,x2,y2=self._bbox(r,c)
        st=self.model.merged_style(r,c)
        fw="bold"   if st["bold"]   else "normal"
        fi="italic" if st["italic"] else "roman"
        fn=(FONT_NAME,st["font_size"],fw,fi)
        
        val = "" if clear else self.model.get(r,c)
        self.entry_var.set(val)
        
        self.inline_entry.configure(font=fn, bg=st["bg"], fg=st["fg"])
        self.canvas.create_window(x1,y1, anchor="nw", window=self.inline_entry,
                                  width=x2-x1, height=y2-y1, tags="inline_entry")
        self.inline_entry.focus_set()
        if not clear:
            self.inline_entry.select_range(0,"end")
            self.inline_entry.icursor("end")

    def _entry_commit(self, e=None):
        if self.editing:
            self.model.save_history()
            r,c=self.editing; v=self.entry_var.get()
            self.model.set(r,c,v); self.formula_var.set(v)
            self.editing=None; self.canvas.delete("inline_entry"); self.render_all()
            self.canvas.focus_set()

    def _entry_cancel(self, e=None):
        self.editing=None; self.canvas.delete("inline_entry"); self.render_all()
        self.canvas.focus_set()

    def _entry_tab(self, e=None):
        self._entry_commit()
        self._move_selection(0, 1)
        return "break"

    def _entry_shifttab(self, e=None):
        self._entry_commit()
        self._move_selection(0, -1)
        return "break"

    def _commit_edit(self):
        if self.editing: self._entry_commit()

    def _formula_commit(self, e=None):
        if self.sel_start:
            self.model.save_history()
            r,c=self.sel_start; self.model.set(r,c,self.formula_var.get()); self.render_all()

    # ── formatação ───────────────────────────
    def _apply(self, fn):
        self.model.save_history()
        for r,c in self._sel_set():
            st=self.model.merged_style(r,c); fn(st); self.model.set_style(r,c,st)
        self.render_all()

    def toggle_bold(self):
        cl=list(self._sel_set())
        if not cl: return
        ab=all(self.model.merged_style(r,c)["bold"] for r,c in cl)
        self._apply(lambda s:s.update({"bold":not ab}))

    def toggle_italic(self):
        cl=list(self._sel_set())
        if not cl: return
        ai=all(self.model.merged_style(r,c)["italic"] for r,c in cl)
        self._apply(lambda s:s.update({"italic":not ai}))

    def pick_bg(self):
        col=colorchooser.askcolor(title="Cor de fundo")[1]
        if col: self._apply(lambda s:s.update({"bg":col}))

    def pick_fg(self):
        col=colorchooser.askcolor(title="Cor da fonte")[1]
        if col: self._apply(lambda s:s.update({"fg":col}))

    def align_left(self):   self._apply(lambda s:s.update({"align":"w"}))
    def align_center(self): self._apply(lambda s:s.update({"align":"center"}))
    def align_right(self):  self._apply(lambda s:s.update({"align":"e"}))
    
    def toggle_all_borders(self):
        cl = list(self._sel_set())
        if not cl: return
        self.model.save_history()
        ab = all(self.model.merged_style(r,c).get("bt") and 
                 self.model.merged_style(r,c).get("bb") and
                 self.model.merged_style(r,c).get("bl") and
                 self.model.merged_style(r,c).get("br") for r,c in cl)
        for r,c in cl:
            mr, mc = self.model.master_of(r, c) or (r,c)
            st = self.model.get_style(mr, mc)
            if ab:
                for b in ("bt","bb","bl","br"): st.pop(b, None)
            else:
                for b in ("bt","bb","bl","br"): st[b] = True
            self.model.set_style(mr, mc, st)
        self.render_all()

    def toggle_borders(self):
        cl = list(self._sel_set())
        if not cl: return
        
        self.model.save_history()
        # Encontra o bounding box da seleção
        min_r = min(r for r,c in cl); max_r = max(r for r,c in cl)
        min_c = min(c for r,c in cl); max_c = max(c for r,c in cl)
        
        # Verifica se deve aplicar ou remover (Toggle) baseado na primeira célula
        st0 = self.model.merged_style(min_r, min_c)
        removing = any(st0.get(b) for b in ("bt","bb","bl","br"))
        
        for r,c in cl:
            mr, mc = self.model.master_of(r, c) or (r,c)
            st = self.model.get_style(mr, mc)
            if removing:
                for b in ("bt","bb","bl","br"): st.pop(b, None)
            else:
                # Aplica as bordas apenas nas extremidades do bloco selecionado
                if r == min_r: st["bt"] = True
                if r == max_r: st["bb"] = True
                if c == min_c: st["bl"] = True
                if c == max_c: st["br"] = True
            self.model.set_style(mr, mc, st)
        self.render_all()

    def font_increase(self):self._apply(lambda s:s.update({"font_size":min(s["font_size"]+1,36)}))
    def font_decrease(self):self._apply(lambda s:s.update({"font_size":max(s["font_size"]-1,6)}))

    # ── mesclar ──────────────────────────────
    def merge_selection(self):
        if not self.sel_start or not self.sel_end: return
        r1,c1=self.sel_start; r2,c2=self.sel_end
        if r1==r2 and c1==c2:
            messagebox.showinfo("Mesclar","Selecione mais de uma célula."); return
        self.model.save_history()
        self.model.merge_range(r1,c1,r2,c2)
        self.sel_start=(min(r1,r2),min(c1,c2)); self.sel_end=self.sel_start
        self.render_all()

    def unmerge_selection(self):
        if not self.sel_start: return
        r1,c1=self.sel_start; r2,c2=self.sel_end or self.sel_start
        self.model.save_history()
        self.model.unmerge_range(r1,c1,r2,c2); self.render_all()

    # ── copiar/colar/recortar ────────────────
    def copy( self):
        sel=self._sel_set()
        if not sel: return
        rs=sorted({r for r,c in sel}); cs=sorted({c for r,c in sel})
        r0,c0=rs[0],cs[0]
        
        # Clipboard interno (com estilo)
        self.clipboard={
            "data":  {(r-r0,c-c0):self.model.get(r,c)       for r,c in sel},
            "styles":{(r-r0,c-c0):self.model.get_style(r,c) for r,c in sel},
        }; self.cut_cells=None
        
        # Clipboard do sistema (texto TSV para Excel)
        txt_rows = []
        for r in range(min(rs), max(rs)+1):
            txt_cols = []
            for c in range(min(cs), max(cs)+1):
                txt_cols.append(str(self.model.get(r,c)))
            txt_rows.append("\t".join(txt_cols))
        tsv = "\n".join(txt_rows)
        self.canvas.clipboard_clear()
        self.canvas.clipboard_append(tsv)

    def cut(self):
        self.copy()
        self.model.save_history()
        for r,c in self._sel_set():
            self.model.set(r,c,"")
            self.model.set_style(r,c,{})
        self.render_all()

    def paste(self):
        if not self.sel_start: return
        self.model.save_history()
        r0, c0 = self.sel_start
        
        # 1. Tenta obter do sistema primeiro (Excel/Outros)
        try:
            sys_text = self.canvas.clipboard_get()
            if sys_text and ("\t" in sys_text or "\n" in sys_text):
                # Se houver tabs ou newlines, assume que é planilha (Excel)
                rows = sys_text.strip("\r\n").split("\n")
                for dr, row_str in enumerate(rows):
                    cols = row_str.split("\t")
                    for dc, val in enumerate(cols):
                        nr, nc = r0+dr, c0+dc
                        if 0<=nr<self.rows and 0<=nc<self.cols:
                            self.model.set(nr, nc, val.strip())
                self.render_all()
                return "break"
            elif sys_text:
                # Se for apenas um texto simples, e não houver clipboard interno multi-celula
                if not self.clipboard or len(self.clipboard["data"]) <= 1:
                    self.model.set(r0, c0, sys_text.strip())
                    self.render_all()
                    return "break"
        except: pass

        # 2. Se falhou sistema ou não é planilha, usa interno (se existir)
        if self.clipboard:
            for (dr,dc),v in self.clipboard["data"].items():
                nr,nc=r0+dr,c0+dc
                if 0<=nr<self.rows and 0<=nc<self.cols: self.model.set(nr,nc,v)
            for (dr,dc),st in self.clipboard["styles"].items():
                nr,nc=r0+dr,c0+dc
                if 0<=nr<self.rows and 0<=nc<self.cols and st: self.model.set_style(nr,nc,st)
            self.render_all()
            return "break"
        
        return "break"

    # ── inserir/excluir linhas e colunas ─────
    def insert_row(self, at):
        self.model.save_history()
        def sk(r,c): return (r+1,c) if r>=at else (r,c)
        self.model.data       ={sk(r,c):v      for (r,c),v in self.model.data.items()}
        self.model.styles     ={sk(r,c):s      for (r,c),s in self.model.styles.items()}
        self.model.merges     ={sk(r,c):sk(*v) for (r,c),v in self.model.merges.items()}
        self.model.row_heights={(r+1 if r>=at else r):h for r,h in self.model.row_heights.items()}
        self.rows=min(self.rows+1,ROWS+50); self.render_all()

    def delete_row(self, at):
        self.model.save_history()
        def sk(r,c): return (r-1,c) if r>at else (r,c)
        self.model.data       ={sk(r,c):v      for (r,c),v in self.model.data.items()   if r!=at}
        self.model.styles     ={sk(r,c):s      for (r,c),s in self.model.styles.items() if r!=at}
        self.model.merges     ={sk(r,c):sk(*v) for (r,c),v in self.model.merges.items() if r!=at}
        self.model.row_heights={(r-1 if r>at else r):h for r,h in self.model.row_heights.items() if r!=at}
        self.render_all()

    def insert_col(self, at):
        self.model.save_history()
        def sk(r,c): return (r,c+1) if c>=at else (r,c)
        self.model.data  ={sk(r,c):v      for (r,c),v in self.model.data.items()}
        self.model.styles={sk(r,c):s      for (r,c),s in self.model.styles.items()}
        self.model.merges={sk(r,c):sk(*v) for (r,c),v in self.model.merges.items()}
        self.model.col_widths={(c+1 if c>=at else c):w for c,w in self.model.col_widths.items()}
        self.cols=min(self.cols+1, COLS+50); self.render_all()

    def delete_col(self, at):
        self.model.save_history()
        def sk(r,c): return (r,c-1) if c>at else (r,c)
        self.model.data  ={sk(r,c):v      for (r,c),v in self.model.data.items()   if c!=at}
        self.model.styles={sk(r,c):s      for (r,c),s in self.model.styles.items() if c!=at}
        self.model.merges={sk(r,c):sk(*v) for (r,c),v in self.model.merges.items() if c!=at}
        self.model.col_widths={(c-1 if c>at else c):w for c,w in self.model.col_widths.items() if c!=at}
        self.render_all()

# ── app ──────────────────────────────────────────────────────────────────────
class ArCondicionadoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1150x720")
        self.resizable(True, True)

        model = SheetModel()
        self.sheet = SpreadsheetWidget(self, model, DEFAULT_SAVE)
        self.sheet.pack(fill="both", expand=True)
        self.sheet.file_mgr.load_json()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        if messagebox.askyesno("Sair", "Deseja salvar antes de sair?"):
            self.sheet.file_mgr.save()
        self.destroy()

if __name__ == "__main__":
    app = ArCondicionadoApp()
    app.mainloop()
